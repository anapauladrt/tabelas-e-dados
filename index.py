import openpyxl as pd
#criar planilha
tabela=pd.Workbook()
#visualizar
print(tabela.sheetnames)
#selecionar pagina
tabela.create_sheet('dados')
dados_page=tabela['dados']
dados_page.append(['Nome', 'Idade' , 'Cidade'])
dados_page.append(['Ana', '19' , 'Canoas'])
dados_page.append(['João', '20' , 'Poa'])
dados_page.append(['Ricardo', '19' , 'Canoas'])
dados_page.append(['José', '40' , 'Poa'])
dados_page.append(['Ana Clara', '19' , 'Gravataí'])
dados_page.append(['João Pedro', '30' , 'Poa'])
dados_page.append(['Rafael', '10' , 'Canoas'])
tabela.save('tabela.xlsx')